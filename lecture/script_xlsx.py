#как искать библиотеки по своим задачам. есть хранилище библиотек https://pypi.org/
#pip install openpyxl

from openpyxl import load_workbook

workbook = load_workbook('../tmp/file_example_XLSX_50.xlsx') #подгружаем наш файл
sheet = workbook.active #active это то, что дает нам работу с листом
print(sheet.cell(row=2, column=3).value) #выводим значение в конкретной ячейке
print(sheet.cell(row=2, column=2).value)
print(sheet.cell(row=1, column=2).value)

for x in sheet.columns: #для каждого x в sheet.columns
    print(x) #использовали принт

#вывести значения ячеек не смогли, препод пошел создавать xls 01:11:50
