import codecs
import csv
import os
from zipfile import ZipFile
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
from conftest import FILE_DIV


def test_check_pdf():
    with ZipFile(os.path.join(FILE_DIV, "new_zip2025.zip"), 'r') as pdf_zip:
        reader_pdf = PdfReader(pdf_zip.open("tmp/Python Testing with Pytest (Brian Okken).pdf"))
        assert len(reader_pdf.pages) == 256  # проверяем кол-во страниц
        assert 'Simple, Rapid, Effective, and Scalable' in reader_pdf.pages[1] ##проверка, что данный текст находится на 1ой страницe
        print("Проверка PDF прошла")

def test_check_xlsx():
    with ZipFile(os.path.join(FILE_DIV, "new_zip2025.zip"), 'r') as xlsx_zip:
        workbook = load_workbook(xlsx_zip.open("tmp/file_example_XLSX_50.xlsx"))
        reader_xlsx = workbook.active
        assert reader_xlsx.cell(row=2, column=2).value == 'Dulce'  # проверяем, что в данной ячейке слово 'Dulce'
        print("Проверка xlsx прошла")

def test_check_csv():
    with ZipFile(os.path.join(FILE_DIV, "new_zip2025.zip"), 'r') as csv_zip:
        reader_csv = csv.reader(codecs.iterdecode(csv_zip.open("tmp/csvtest.csv"), 'utf-8'), delimiter=';')
        rows = list(reader_csv)
        assert any('Andrey' in row for row in rows) #ищем текст Андрей по всем строка
        assert len(rows) == 4 #проверяет количество строк
        print("Проверка csv прошла")
